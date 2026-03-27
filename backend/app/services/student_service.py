"""Student business logic service."""

import io
import uuid
from pathlib import Path

from fastapi import UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.error_codes import ErrorCode
from app.core.exceptions import DuplicateException, NotFoundException, ValidationException
from app.models.category import Category
from app.models.student import Student
from app.repositories.student_repo import StudentRepository
from app.schemas.student import (
    StudentCreate,
    StudentImportResult,
    StudentUpdate,
)

ALLOWED_FACE_TYPES = {"image/jpeg", "image/png"}
MAX_FACE_SIZE = 5 * 1024 * 1024  # 5MB


class StudentService:
    def __init__(self, session: AsyncSession):
        self.repo = StudentRepository(session)
        self.session = session

    async def _resolve_category_name(self, category_id: int) -> str | None:
        """Return category name for syncing class_name."""
        from sqlalchemy import select
        result = await self.session.execute(
            select(Category).where(Category.id == category_id)
        )
        cat = result.scalar_one_or_none()
        return cat.name if cat else None

    async def create_student(self, data: StudentCreate) -> Student:
        if data.external_id:
            existing = await self.repo.get_by_external_id(data.external_id)
            if existing:
                raise DuplicateException(
                    f"Student with external_id '{data.external_id}' already exists.",
                    error_code=ErrorCode.STUDENT_DUPLICATE_EXTERNAL_ID,
                )
        if data.employee_no:
            existing = await self.repo.get_by_employee_no(data.employee_no)
            if existing:
                raise DuplicateException(
                    f"Student with employee_no '{data.employee_no}' already exists.",
                    error_code=ErrorCode.STUDENT_DUPLICATE_EMPLOYEE_NO,
                )
        create_data = data.model_dump(exclude_unset=True)
        # Sync class_name from category
        if data.category_id:
            cat_name = await self._resolve_category_name(data.category_id)
            if cat_name:
                create_data["class_name"] = cat_name
        return await self.repo.create(create_data)

    async def get_student(self, student_id: uuid.UUID) -> Student:
        student = await self.repo.get(student_id)
        if not student:
            raise NotFoundException(
                f"Student with id '{student_id}' not found.",
                error_code=ErrorCode.STUDENT_NOT_FOUND,
            )
        return student

    async def list_students(
        self,
        skip: int = 0,
        limit: int = 100,
        class_name: str | None = None,
        search: str | None = None,
        is_active: bool = True,
        sort: str = "-created_at",
        category_id: int | None = None,
        no_category: bool = False,
    ) -> tuple[list[Student], int]:
        return await self.repo.get_all_students(
            skip, limit, class_name, search, is_active, sort, category_id,
            no_category=no_category,
        )

    async def update_student(
        self, student_id: uuid.UUID, data: StudentUpdate
    ) -> Student:
        update_data = data.model_dump(exclude_unset=True)
        if not update_data:
            return await self.get_student(student_id)

        # Check uniqueness if changing external_id or employee_no
        if "external_id" in update_data and update_data["external_id"]:
            existing = await self.repo.get_by_external_id(update_data["external_id"])
            if existing and existing.id != student_id:
                raise DuplicateException(
                    f"Student with external_id '{update_data['external_id']}' already exists.",
                    error_code=ErrorCode.STUDENT_DUPLICATE_EXTERNAL_ID,
                )
        if "employee_no" in update_data and update_data["employee_no"]:
            existing = await self.repo.get_by_employee_no(update_data["employee_no"])
            if existing and existing.id != student_id:
                raise DuplicateException(
                    f"Student with employee_no '{update_data['employee_no']}' already exists.",
                    error_code=ErrorCode.STUDENT_DUPLICATE_EMPLOYEE_NO,
                )

        # Sync class_name from category if category_id is being set
        if "category_id" in update_data and update_data["category_id"]:
            cat_name = await self._resolve_category_name(update_data["category_id"])
            if cat_name:
                update_data["class_name"] = cat_name

        return await self.repo.update(student_id, update_data)

    async def delete_student(self, student_id: uuid.UUID) -> Student:
        return await self.repo.soft_delete(student_id)

    async def register_face(
        self, student_id: uuid.UUID, image: UploadFile
    ) -> Student:
        """Save face image and mark face_registered=True."""
        student = await self.get_student(student_id)

        # Validate file type
        content_type = image.content_type or ""
        if content_type not in ALLOWED_FACE_TYPES:
            raise ValidationException(
                "Invalid image format. Only JPEG and PNG are allowed.",
                error_code=ErrorCode.INVALID_FACE_IMAGE,
            )

        # Read and validate size
        content = await image.read()
        if len(content) > MAX_FACE_SIZE:
            raise ValidationException(
                "Image file too large. Maximum size is 5MB.",
                error_code=ErrorCode.INVALID_FACE_IMAGE,
            )

        upload_dir = Path("./data/faces")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / f"{student_id}.jpg"

        # Convert PNG to JPEG if needed
        if content_type == "image/png":
            try:
                from PIL import Image

                img = Image.open(io.BytesIO(content))
                buf = io.BytesIO()
                img.convert("RGB").save(buf, format="JPEG", quality=90)
                content = buf.getvalue()
            except Exception:
                pass  # Keep original if conversion fails

        file_path.write_bytes(content)

        student.face_image_path = str(file_path)
        student.face_registered = True
        await self.session.flush()
        await self.session.refresh(student)
        return student

    async def import_from_excel(
        self, file: UploadFile
    ) -> StudentImportResult:
        """Parse an Excel file and bulk-create/update students."""
        filename = file.filename or ""
        if not filename.endswith(".xlsx"):
            raise ValidationException(
                "Only .xlsx files are supported.",
                error_code=ErrorCode.IMPORT_VALIDATION_ERROR,
            )

        import openpyxl

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise ValidationException(
                "Excel file has no active sheet.",
                error_code=ErrorCode.IMPORT_VALIDATION_ERROR,
            )

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return StudentImportResult(total=0, created=0, updated=0, errors=[])

        # Parse header
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        required = {"name", "class_name"}
        if not required.issubset(set(header)):
            raise ValidationException(
                f"Missing required columns: {required - set(header)}",
                error_code=ErrorCode.IMPORT_VALIDATION_ERROR,
            )

        created = 0
        updated = 0
        errors: list[str] = []
        total = len(rows) - 1

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                row_data = dict(zip(header, row, strict=False))
                name = str(row_data.get("name", "")).strip()
                if not name:
                    errors.append(f"Row {row_idx}: missing name")
                    continue

                class_name = str(row_data.get("class_name", "")).strip() or None
                external_id = str(row_data.get("external_id", "")).strip() or None
                employee_no = str(row_data.get("employee_no", "")).strip() or None
                phone = str(row_data.get("phone", "")).strip() or None
                parent_phone = str(row_data.get("parent_phone", "")).strip() or None

                # Auto-generate employee_no if not provided
                if not employee_no:
                    count = await self.repo.count({})
                    employee_no = f"ATX-{count + row_idx:06d}"

                # Upsert by external_id
                if external_id:
                    existing = await self.repo.get_by_external_id(external_id)
                    if existing:
                        await self.repo.update(existing.id, {
                            "name": name,
                            "class_name": class_name,
                            "phone": phone,
                            "parent_phone": parent_phone,
                            "employee_no": employee_no,
                        })
                        updated += 1
                        continue

                await self.repo.create({
                    "name": name,
                    "class_name": class_name,
                    "external_id": external_id,
                    "employee_no": employee_no,
                    "phone": phone,
                    "parent_phone": parent_phone,
                })
                created += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {e}")

        wb.close()
        return StudentImportResult(
            total=total, created=created, updated=updated, errors=errors
        )

    async def export_to_excel(
        self, class_name: str | None = None
    ) -> bytes:
        """Generate an XLSX export of students."""
        import openpyxl

        students, _ = await self.repo.get_all_students(
            skip=0, limit=10000, class_name=class_name
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = [
            "Name", "Class", "Phone", "Parent Phone",
            "External ID", "Employee No", "Face Registered", "Active",
        ]
        ws.append(headers)

        for s in students:
            ws.append([
                s.name,
                s.class_name or "",
                s.phone or "",
                s.parent_phone or "",
                s.external_id or "",
                s.employee_no or "",
                "Yes" if s.face_registered else "No",
                "Yes" if s.is_active else "No",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def export_to_csv(
        self, class_name: str | None = None
    ) -> bytes:
        """Generate a CSV export of students."""
        import csv

        students, _ = await self.repo.get_all_students(
            skip=0, limit=10000, class_name=class_name
        )

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "Name", "Class", "Phone", "Parent Phone",
            "External ID", "Employee No", "Face Registered", "Active",
        ])

        for s in students:
            writer.writerow([
                s.name,
                s.class_name or "",
                s.phone or "",
                s.parent_phone or "",
                s.external_id or "",
                s.employee_no or "",
                "Yes" if s.face_registered else "No",
                "Yes" if s.is_active else "No",
            ])

        return buf.getvalue().encode("utf-8")
